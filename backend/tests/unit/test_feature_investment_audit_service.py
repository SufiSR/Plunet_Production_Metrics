from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import Session, sessionmaker

import app.models  # noqa: F401
from app.jira_analytics.feature_investment_audit_service import (
    feature_investment_audit,
    feature_investment_audit_issues,
    feature_investment_audit_worklogs,
    feature_investment_audit_xlsx,
)
from app.jira_analytics.models import (
    JiraFeatureFamily,
    JiraFeatureFamilyMember,
    JiraFeatureRoot,
    JiraIssue,
    JiraIssueDetail,
    JiraProject,
    JiraUser,
    JiraUserMonthlyHrworksHours,
    JiraWorklog,
    MonthlyAllocatedEffort,
    MonthlyTopicEffortBase,
)
from app.models.base import Base


def _session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)

    @event.listens_for(engine, "connect")
    def _enable_fk(dbapi_conn, _rec):  # type: ignore[no-untyped-def]
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, future=True)
    return factory()


def _issue(
    db: Session,
    project: JiraProject,
    *,
    key: str,
    summary: str,
    issue_type: str = "Task",
) -> JiraIssue:
    issue = JiraIssue(
        jira_issue_id=f"id-{key}",
        key=key,
        project_id=project.id,
        issue_type_name=issue_type,
        summary=summary,
        last_seen_at=datetime.now(timezone.utc),
    )
    db.add(issue)
    db.flush()
    return issue


def _feature(
    db: Session,
    project: JiraProject,
    *,
    key: str,
    name: str,
) -> tuple[JiraFeatureRoot, JiraIssue]:
    issue = _issue(db, project, key=key, summary=name, issue_type="Feature")
    root = JiraFeatureRoot(
        root_issue_id=issue.id,
        root_key=key,
        root_project_key=project.key,
        root_issue_type_name="Feature",
        name=name,
        detection_rule="unit-test",
    )
    db.add(root)
    db.flush()
    return root, issue


def test_feature_investment_audit_scales_feature_share_and_keeps_hrworks_audit() -> None:
    with _session() as db:
        project = JiraProject(jira_project_id="100", key="PMGT", name="PMGT")
        db.add(project)
        db.flush()
        family = JiraFeatureFamily(name="Audit Family", active=True)
        db.add(family)
        dev = JiraUser(account_id="dev-1", email_address="dev@example.com", display_name="Dev One")
        po = JiraUser(account_id="po-1", email_address="po@example.com", display_name="PO")
        db.add_all([dev, po])
        db.flush()
        root, _feature_issue = _feature(db, project, key="PMGT-1", name="Feature A")
        db.add(
            JiraIssueDetail(
                issue_id=root.root_issue_id,
                start_date=date(2026, 5, 2),
                promised_delivery_date=date(2026, 7, 15),
                delivery_status="In Progress",
            )
        )
        db.add(JiraFeatureFamilyMember(family_id=family.id, feature_root_id=root.id))
        feature_issue = _issue(db, project, key="BM-1", summary="Feature work")
        support_issue = _issue(db, project, key="SUP-1", summary="Support work")
        db.add_all(
            [
                JiraUserMonthlyHrworksHours(
                    jira_user_id=dev.id,
                    month_start=date(2026, 5, 1),
                    month_end=date(2026, 5, 31),
                    planned_working_hours=Decimal("160"),
                    clocked_working_hours=Decimal("150"),
                ),
                JiraUserMonthlyHrworksHours(
                    jira_user_id=po.id,
                    month_start=date(2026, 5, 1),
                    month_end=date(2026, 5, 31),
                    planned_working_hours=Decimal("100"),
                    clocked_working_hours=Decimal("95"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 5, 1),
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=feature_issue.id,
                    issue_key=feature_issue.key,
                    issue_type_name=feature_issue.issue_type_name,
                    summary=feature_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id=dev.account_id,
                    display_name=dev.display_name,
                    role_name="Developer",
                    topic_type="feature",
                    direct_hours=Decimal("10"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 5, 1),
                    issue_id=support_issue.id,
                    issue_key=support_issue.key,
                    issue_type_name=support_issue.issue_type_name,
                    summary=support_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id=dev.account_id,
                    display_name=dev.display_name,
                    role_name="Developer",
                    topic_type="tech_support",
                    direct_hours=Decimal("30"),
                ),
                MonthlyAllocatedEffort(
                    period_month=date(2026, 5, 1),
                    topic_type="feature",
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=feature_issue.id,
                    issue_key=feature_issue.key,
                    team_name="Team Tantrum",
                    source_user_email="po@example.com",
                    source_display_name="PO",
                    source_role_name="Product Owner",
                    allocation_kind="indirect_allocated",
                    hours=Decimal("5"),
                    rule_snapshot_json={},
                ),
                MonthlyAllocatedEffort(
                    period_month=date(2026, 5, 1),
                    topic_type="tech_support",
                    issue_id=support_issue.id,
                    issue_key=support_issue.key,
                    team_name="Team Tantrum",
                    source_user_email="po@example.com",
                    source_display_name="PO",
                    source_role_name="Product Owner",
                    allocation_kind="indirect_allocated",
                    hours=Decimal("7"),
                    rule_snapshot_json={},
                ),
                JiraWorklog(
                    issue_id=feature_issue.id,
                    jira_worklog_id="wl-1",
                    author_user_id=dev.id,
                    author_account_id=dev.account_id,
                    author_display_name=dev.display_name,
                    started_at=datetime(2026, 5, 10, tzinfo=timezone.utc),
                    time_spent_seconds=36000,
                ),
            ]
        )
        db.commit()

        report = feature_investment_audit(
            db,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 31),
        )
        issues = feature_investment_audit_issues(
            db,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 31),
            feature_key="PMGT-1",
        )
        worklogs = feature_investment_audit_worklogs(
            db,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 31),
            issue_key="BM-1",
        )
        workbook = load_workbook(
            BytesIO(
                feature_investment_audit_xlsx(
                    db,
                    date_from=date(2026, 5, 1),
                    date_to=date(2026, 5, 31),
                )
            )
        )

    rows_by_feature = {row["feature_identifier"]: row for row in report.table}
    row = rows_by_feature["PMGT-1"]
    support_row = rows_by_feature["__other_misc__"]
    assert row["booked_hours"] == 10.0
    assert row["calculated_hours"] == 37.0
    assert row["overhead_hours"] == 5.0
    assert support_row["feature_name"] == "Other misc"
    assert support_row["booked_hours"] == 30.0
    assert support_row["calculated_hours"] == 103.0
    assert support_row["overhead_hours"] == 7.0
    assert report.summary["booked_hours"] == 40.0
    assert report.summary["calculated_hours"] == 140.0
    assert issues.table[0]["issue_identifier"] == "BM-1"
    assert issues.table[0]["calculated_hours"] == 37.0
    direct_worklog = next(item for item in worklogs.table if item["source"] == "Jira worklog")
    assert direct_worklog["scale_factor"] == 3.2
    assert direct_worklog["hrworks_planned_hours"] == 160.0
    assert workbook.sheetnames == [
        "Calculated summary",
        "Methodik & Definitionen",
        "Actual summary",
        "Issue detail",
        "Feature by month and role",
        "Family by month and role",
        "HRWorks audit",
    ]
    assert workbook["Calculated summary"]["A1"].value == "Feature Family"
    assert workbook["Calculated summary"]["D1"].value == "2026-05 Calculated"
    assert workbook["Calculated summary"]["E1"].value == "2026-05 Overhead"
    assert workbook["Calculated summary"]["F1"].value == "2026-05 Calc Total"
    assert workbook["Calculated summary"]["B2"].value == "Other misc"
    assert workbook["Calculated summary"]["D2"].value == 96.0
    assert workbook["Calculated summary"]["E2"].value == 7.0
    assert workbook["Calculated summary"]["F2"].value == 103.0
    assert workbook["Calculated summary"]["B3"].value == "Feature A"
    assert workbook["Calculated summary"]["D3"].value == 32.0
    assert workbook["Calculated summary"]["E3"].value == 5.0
    assert workbook["Calculated summary"]["F3"].value == 37.0
    assert workbook["Actual summary"]["B2"].value == "Other misc"
    assert workbook["Actual summary"]["D2"].value == 30.0
    assert workbook["Actual summary"]["E2"].value == 7.0
    assert workbook["Actual summary"]["F2"].value == 37.0
    assert workbook["Actual summary"]["B3"].value == "Feature A"
    assert workbook["Actual summary"]["D3"].value == 10.0
    assert workbook["Actual summary"]["E3"].value == 5.0
    assert workbook["Actual summary"]["F3"].value == 15.0
    assert workbook["Issue detail"]["C1"].value == "Feature family start date"
    assert workbook["Issue detail"]["C2"].value == "2026-05-02"
    assert workbook["Issue detail"]["D2"].value == "2026-07-15"
    assert workbook["Issue detail"]["E2"].value == "In Progress"
    assert workbook["Issue detail"]["H2"].value == "2026-05-02"
    assert workbook["Issue detail"]["I2"].value == "2026-07-15"
    assert workbook["Issue detail"]["J2"].value == "In Progress"
    assert workbook["Feature by month and role"]["C2"].value == "2026-05-02"
    assert workbook["Family by month and role"]["C2"].value == "2026-05-02"
    assert workbook["HRWorks audit"]["E2"].value == 160.0
    assert workbook["HRWorks audit"]["F2"].value == 40.0
    assert workbook["HRWorks audit"]["A3"].value == "PO"
    assert workbook["HRWorks audit"]["E3"].value == 100.0
    assert workbook["HRWorks audit"]["F3"].value == 0.0
    assert workbook["HRWorks audit"]["G3"].value == 0.0
    assert workbook["HRWorks audit"]["H3"].value == 1.0
    assert workbook["HRWorks audit"]["I3"].value is False
    assert workbook["HRWorks audit"]["J3"].value == 12.0
    assert workbook["HRWorks audit"]["K3"].value == "Allocated overhead"
    methodology = workbook["Methodik & Definitionen"]
    assert methodology["A1"].value == "Feature Investment Audit - Methodik & Definitionen"
    assert methodology["A2"].value == "Bereich"
    assert methodology["B2"].value == "Beschreibung"
    assert methodology.freeze_panes == "A3"
    assert methodology.sheet_properties.tabColor.rgb.endswith("806000")
    assert methodology.column_dimensions["A"].width == 34
    assert methodology.column_dimensions["B"].width == 110
    assert methodology["B3"].alignment.wrap_text is True
    assert "2026-05" in methodology["B3"].value
    assert "MonthlyTopicEffortBase" in methodology["B4"].value
    assert "Gebuchte Stunden (Booked / Actual)" in methodology["B8"].value
    assert "HRWorks planned hours * 80%" in methodology["B8"].value
    assert "Product Owner" in methodology["B8"].value


def test_feature_investment_audit_xlsx_summary_months_latest_first_and_sorted() -> None:
    with _session() as db:
        project = JiraProject(jira_project_id="100", key="PMGT", name="PMGT")
        db.add(project)
        db.flush()
        family = JiraFeatureFamily(name="Sort Family", active=True)
        db.add(family)
        db.flush()
        ux = JiraUser(account_id="ux-1", email_address="ux@example.com", display_name="UX One")
        db.add(ux)
        db.flush()
        root, _feature_issue = _feature(db, project, key="PMGT-10", name="Feature Sort")
        db.add(JiraFeatureFamilyMember(family_id=family.id, feature_root_id=root.id))
        feature_issue = _issue(db, project, key="BM-10", summary="Feature detail")
        support_issue = _issue(db, project, key="SUP-10", summary="Support detail")
        db.add_all(
            [
                JiraUserMonthlyHrworksHours(
                    jira_user_id=ux.id,
                    month_start=date(2026, 5, 1),
                    month_end=date(2026, 5, 31),
                    planned_working_hours=Decimal("150"),
                    clocked_working_hours=Decimal("145"),
                ),
                JiraUserMonthlyHrworksHours(
                    jira_user_id=ux.id,
                    month_start=date(2026, 6, 1),
                    month_end=date(2026, 6, 30),
                    planned_working_hours=Decimal("160"),
                    clocked_working_hours=Decimal("155"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 5, 1),
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=feature_issue.id,
                    issue_key=feature_issue.key,
                    issue_type_name=feature_issue.issue_type_name,
                    summary=feature_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="feature",
                    direct_hours=Decimal("100"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 6, 1),
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=feature_issue.id,
                    issue_key=feature_issue.key,
                    issue_type_name=feature_issue.issue_type_name,
                    summary=feature_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="feature",
                    direct_hours=Decimal("2"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 5, 1),
                    issue_id=support_issue.id,
                    issue_key=support_issue.key,
                    issue_type_name=support_issue.issue_type_name,
                    summary=support_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="tech_support",
                    direct_hours=Decimal("5"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 6, 1),
                    issue_id=support_issue.id,
                    issue_key=support_issue.key,
                    issue_type_name=support_issue.issue_type_name,
                    summary=support_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="tech_support",
                    direct_hours=Decimal("20"),
                ),
                MonthlyAllocatedEffort(
                    period_month=date(2026, 6, 1),
                    topic_type="feature",
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=feature_issue.id,
                    issue_key=feature_issue.key,
                    team_name="Team Tantrum",
                    source_user_email="po@example.com",
                    source_display_name="PO",
                    source_role_name="Product Owner",
                    allocation_kind="indirect_allocated",
                    hours=Decimal("1"),
                    rule_snapshot_json={},
                ),
                MonthlyAllocatedEffort(
                    period_month=date(2026, 6, 1),
                    topic_type="tech_support",
                    issue_id=support_issue.id,
                    issue_key=support_issue.key,
                    team_name="Team Tantrum",
                    source_user_email="po@example.com",
                    source_display_name="PO",
                    source_role_name="Product Owner",
                    allocation_kind="indirect_allocated",
                    hours=Decimal("2"),
                    rule_snapshot_json={},
                ),
            ]
        )
        db.commit()
        workbook = load_workbook(
            BytesIO(
                feature_investment_audit_xlsx(
                    db,
                    date_from=date(2026, 5, 1),
                    date_to=date(2026, 6, 30),
                )
            )
        )

    calculated = workbook["Calculated summary"]
    assert calculated["D1"].value == "2026-06 Calculated"
    assert calculated["F1"].value == "2026-06 Calc Total"
    assert calculated["G1"].value == "2026-05 Calculated"
    assert calculated["I1"].value == "2026-05 Calc Total"
    assert calculated["B2"].value == "Other misc"
    assert calculated["D2"].value == 20.0
    assert calculated["E2"].value == 2.0
    assert calculated["F2"].value == 22.0
    assert calculated["I2"].value == 5.0
    assert calculated["B3"].value == "Feature Sort"
    assert calculated["D3"].value == 2.0
    assert calculated["E3"].value == 1.0
    assert calculated["F3"].value == 3.0
    assert calculated["I3"].value == 100.0
    actual = workbook["Actual summary"]
    assert actual["D1"].value == "2026-06 Actual"
    assert actual["F1"].value == "2026-06 Actual Total"
    assert actual["G1"].value == "2026-05 Actual"
    assert actual["I1"].value == "2026-05 Actual Total"
    assert actual["B2"].value == "Other misc"
    assert actual["F2"].value == 22.0
    assert actual["B3"].value == "Feature Sort"
    assert actual["F3"].value == 3.0
    audit = workbook["HRWorks audit"]
    ux_rows = [
        row
        for row in audit.iter_rows(min_row=2, values_only=True)
        if row[0] == "UX One"
    ]
    assert [
        (row[3], row[4], row[5], row[7], row[8], row[9], row[10])
        for row in ux_rows
    ] == [
        ("2026-05-01", 150.0, 105.0, 1.0, False, 0.0, "Jira worklog"),
        ("2026-06-01", 160.0, 22.0, 1.0, False, 0.0, "Jira worklog"),
    ]


def test_feature_investment_audit_defaults_to_last_completed_month() -> None:
    with _session() as db:
        project = JiraProject(jira_project_id="100", key="PMGT", name="PMGT")
        db.add(project)
        db.flush()
        may_issue = _issue(db, project, key="UX-1", summary="May work")
        june_issue = _issue(db, project, key="UX-2", summary="Current month work")
        db.add_all(
            [
                MonthlyTopicEffortBase(
                    period_month=date(2026, 5, 1),
                    issue_id=may_issue.id,
                    issue_key=may_issue.key,
                    issue_type_name=may_issue.issue_type_name,
                    summary=may_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="tech_support",
                    direct_hours=Decimal("4"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 6, 1),
                    issue_id=june_issue.id,
                    issue_key=june_issue.key,
                    issue_type_name=june_issue.issue_type_name,
                    summary=june_issue.summary,
                    team_name="Team Tantrum",
                    user_account_id="ux-1",
                    display_name="UX One",
                    role_name="UX",
                    topic_type="tech_support",
                    direct_hours=Decimal("99"),
                ),
            ]
        )
        db.commit()
        with patch("app.jira_analytics.feature_investment_audit_service.datetime") as frozen:
            frozen.now.return_value = datetime(2026, 6, 15, tzinfo=timezone.utc)
            report = feature_investment_audit(
                db,
                date_from=date(2026, 5, 1),
                date_to=None,
            )

    assert report.filters["periods"] == ["2026-05-01"]
    assert report.filters["to"] == "2026-05-31"
    assert report.summary["booked_hours"] == 4.0


def test_feature_investment_audit_never_scales_dev_qa_below_booked_hours() -> None:
    with _session() as db:
        project = JiraProject(jira_project_id="100", key="PMGT", name="PMGT")
        db.add(project)
        db.flush()
        family = JiraFeatureFamily(name="Clamp Family", active=True)
        db.add(family)
        dev = JiraUser(
            account_id="dev-overbooked",
            email_address="over@example.com",
            display_name="Over Booked",
        )
        db.add(dev)
        db.flush()
        root, _feature_issue = _feature(db, project, key="PMGT-2", name="Overbooked Feature")
        db.add(JiraFeatureFamilyMember(family_id=family.id, feature_root_id=root.id))
        issue = _issue(db, project, key="BM-2", summary="Feature work")
        db.add_all(
            [
                JiraUserMonthlyHrworksHours(
                    jira_user_id=dev.id,
                    month_start=date(2026, 4, 1),
                    month_end=date(2026, 4, 30),
                    planned_working_hours=Decimal("152"),
                    clocked_working_hours=Decimal("152"),
                ),
                MonthlyTopicEffortBase(
                    period_month=date(2026, 4, 1),
                    feature_root_id=root.id,
                    feature_key=root.root_key,
                    feature_name=root.name,
                    issue_id=issue.id,
                    issue_key=issue.key,
                    issue_type_name=issue.issue_type_name,
                    summary=issue.summary,
                    team_name="Team Tantrum",
                    user_account_id=dev.account_id,
                    display_name=dev.display_name,
                    role_name="Developer",
                    topic_type="feature",
                    direct_hours=Decimal("157.27"),
                ),
            ]
        )
        db.commit()

        report = feature_investment_audit(
            db,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 30),
        )
        workbook = load_workbook(
            BytesIO(
                feature_investment_audit_xlsx(
                    db,
                    date_from=date(2026, 4, 1),
                    date_to=date(2026, 4, 30),
                )
            )
        )

    row = report.table[0]
    assert row["booked_hours"] == 157.27
    assert row["calculated_hours"] == 157.27
    assert workbook["HRWorks audit"]["G2"].value == 121.6
    assert workbook["HRWorks audit"]["H2"].value == 1.0
