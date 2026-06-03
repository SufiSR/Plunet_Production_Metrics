from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.jira_analytics.models import (
    JiraFeatureFamily,
    JiraFeatureFamilyMember,
    JiraFeatureRoot,
    JiraIssue,
    JiraIssueDetail,
    JiraUser,
    JiraUserMonthlyHrworksHours,
    JiraWorklog,
    MonthlyAllocatedEffort,
    MonthlyTopicEffortBase,
)
from app.jira_analytics.project_scope import (
    apply_worklog_issue_scope,
    scope_monthly_allocated_effort,
    scope_monthly_topic_effort,
)
from app.jira_analytics.team_names import normalized_team_name
from app.schemas.jira_analytics_reports import AnalyticsReportResponse

DEV_QA_CAPACITY_FACTOR = 0.8
UNASSIGNED_FAMILY_IDENTIFIER = "__unassigned__"
UNASSIGNED_FAMILY_NAME = "Unassigned feature family"
OTHER_BUG_TYPES = frozenset({"Bug", "Problem"})
OTHER_FEATURE_TYPES = frozenset(
    {
        "Epic",
        "New Feature",
        "Improvement",
        "Feature",
        "Idea",
        "Sub-task",
        "Subtask",
        "Development Subtask",
        "Improvement Subtask",
    }
)
GENERIC_BUCKETS: dict[str, tuple[str, str]] = {
    "other_bug": ("__other_bug__", "Other bug"),
    "other_feature": ("__other_feature__", "Other feature"),
    "other_misc": ("__other_misc__", "Other misc"),
}


@dataclass(frozen=True, slots=True)
class _FamilyInfo:
    identifier: str
    name: str


@dataclass(frozen=True, slots=True)
class _FeatureMeta:
    start_date: str | None
    end_date: str | None
    status: str | None


@dataclass(frozen=True, slots=True)
class _FactorInfo:
    person_key: str
    person: str
    account_id: str | None
    jira_user_id: int | None
    role: str
    role_bucket: str
    team: str | None
    period: date
    planned_hours: float
    direct_booked_hours: float
    capacity_target_hours: float
    scale_factor: float
    missing_hrworks: bool


@dataclass(frozen=True, slots=True)
class _AuditRow:
    period: date
    family_identifier: str
    family_name: str
    family_start_date: str | None
    family_end_date: str | None
    family_status: str | None
    feature_identifier: str
    feature_name: str
    feature_start_date: str | None
    feature_end_date: str | None
    feature_status: str | None
    team: str | None
    issue_identifier: str
    issue_name: str | None
    issue_type: str | None
    person_key: str
    person: str
    role: str
    role_bucket: str
    source: str
    booked_hours: float
    calculated_hours: float
    overhead_hours: float
    scale_factor: float
    hrworks_planned_hours: float


def _monthly_period_bounds(date_from: date | None, date_to: date | None) -> tuple[date, date]:
    today = datetime.now(timezone.utc).date()
    if date_to:
        period_to = date(date_to.year, date_to.month, 1)
    else:
        month = today.month - 1
        year = today.year
        if month <= 0:
            month = 12
            year -= 1
        period_to = date(year, month, 1)
    if date_from:
        period_from = date(date_from.year, date_from.month, 1)
    else:
        month = period_to.month - 5
        year = period_to.year
        while month <= 0:
            month += 12
            year -= 1
        period_from = date(year, month, 1)
    return period_from, period_to


def _month_end(period: date) -> date:
    return _next_month_start(period) - timedelta(days=1)


def _next_month_start(period: date) -> date:
    if period.month == 12:
        return date(period.year + 1, 1, 1)
    return date(period.year, period.month + 1, 1)


def _month_starts(period_from: date, period_to: date) -> list[date]:
    months: list[date] = []
    current = date(period_from.year, period_from.month, 1)
    while current <= period_to:
        months.append(current)
        current = _next_month_start(current)
    return months


def _role_bucket(role_name: str | None) -> str:
    role = (role_name or "").strip().lower()
    if "ux" in role or "designer" in role or "user experience" in role:
        return "direct_ux"
    if "qa" in role or "test" in role or "quality assurance" in role:
        return "direct_qa"
    if "developer" in role or "engineer" in role or role in {"dev", "software engineer"}:
        return "direct_dev"
    if role in {"po", "pm"} or "product owner" in role or "product manager" in role:
        return "product_overhead"
    if "architect" in role or "head of dev" in role or "head of development" in role:
        return "dev_overhead"
    return "other"


def _is_dev_qa_bucket(role_bucket: str) -> bool:
    return role_bucket in {"direct_dev", "direct_qa"}


def _family_map(db: Session) -> dict[int, _FamilyInfo]:
    rows = db.execute(
        select(
            JiraFeatureFamilyMember.feature_root_id,
            JiraFeatureFamily.id,
            JiraFeatureFamily.name,
        )
        .join(JiraFeatureFamily, JiraFeatureFamily.id == JiraFeatureFamilyMember.family_id)
        .where(JiraFeatureFamily.active.is_(True))
    ).all()
    return {
        int(feature_root_id): _FamilyInfo(str(family_id), str(name or family_id))
        for feature_root_id, family_id, name in rows
    }


def _iso_date(value: date | datetime | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value.isoformat()


def _feature_meta_map(db: Session) -> dict[int, _FeatureMeta]:
    rows = db.execute(
        select(
            JiraFeatureRoot.id,
            JiraIssueDetail.start_date,
            JiraIssueDetail.promised_delivery_date,
            JiraIssueDetail.actual_end,
            JiraIssueDetail.delivery_status,
            JiraIssue.status_name,
        )
        .join(JiraIssue, JiraIssue.id == JiraFeatureRoot.root_issue_id)
        .outerjoin(JiraIssueDetail, JiraIssueDetail.issue_id == JiraIssue.id)
        .where(JiraFeatureRoot.active.is_(True))
    ).all()
    return {
        int(root_id): _FeatureMeta(
            start_date=_iso_date(start_date),
            end_date=_iso_date(actual_end) or _iso_date(promised_delivery_date),
            status=(delivery_status or status_name or "").strip() or None,
        )
        for (
            root_id,
            start_date,
            promised_delivery_date,
            actual_end,
            delivery_status,
            status_name,
        ) in rows
    }


def _rollup_date(values: list[str | None], *, latest: bool = False) -> str | None:
    dates = sorted(value for value in values if value)
    if not dates:
        return None
    return dates[-1] if latest else dates[0]


def _rollup_status(values: list[str | None]) -> str | None:
    statuses = sorted({value.strip() for value in values if value and value.strip()})
    if not statuses:
        return None
    if len(statuses) == 1:
        return statuses[0]
    return "Mixed"


def _family_meta_map(
    db: Session,
    *,
    feature_meta: dict[int, _FeatureMeta],
) -> dict[str, _FeatureMeta]:
    rows = db.execute(
        select(JiraFeatureFamilyMember.family_id, JiraFeatureFamilyMember.feature_root_id)
    ).all()
    grouped: dict[str, list[_FeatureMeta]] = defaultdict(list)
    for family_id, feature_root_id in rows:
        meta = feature_meta.get(int(feature_root_id))
        if meta is not None:
            grouped[str(family_id)].append(meta)
    return {
        family_id: _FeatureMeta(
            start_date=_rollup_date([meta.start_date for meta in metas]),
            end_date=_rollup_date([meta.end_date for meta in metas], latest=True),
            status=_rollup_status([meta.status for meta in metas]),
        )
        for family_id, metas in grouped.items()
    }


def _feature_dimensions(
    *,
    feature_root_id: int | None,
    feature_key: str | None,
    feature_name: str | None,
    families: dict[int, _FamilyInfo],
) -> tuple[str, str, str, str]:
    family = families.get(feature_root_id or -1)
    return (
        family.identifier if family else UNASSIGNED_FAMILY_IDENTIFIER,
        family.name if family else UNASSIGNED_FAMILY_NAME,
        str(feature_key or "unknown"),
        str(feature_name or feature_key or "Unknown feature"),
    )


def _metadata_dimensions(
    *,
    feature_root_id: int | None,
    family_identifier: str,
    feature_meta: dict[int, _FeatureMeta],
    family_meta: dict[str, _FeatureMeta],
) -> tuple[str | None, str | None, str | None, str | None, str | None, str | None]:
    feature = feature_meta.get(feature_root_id or -1)
    family = family_meta.get(family_identifier) or feature
    return (
        family.start_date if family else None,
        family.end_date if family else None,
        family.status if family else None,
        feature.start_date if feature else None,
        feature.end_date if feature else None,
        feature.status if feature else None,
    )


def _generic_bucket(topic_type: str | None, issue_type: str | None) -> tuple[str, str]:
    topic = (topic_type or "").strip()
    if topic == "unassigned_bug":
        return GENERIC_BUCKETS["other_bug"]
    if topic == "issue_without_feature":
        return GENERIC_BUCKETS["other_feature"]
    if topic in {"tech_support", "unclassified"}:
        return GENERIC_BUCKETS["other_misc"]
    issue_type_name = (issue_type or "").strip()
    if issue_type_name in OTHER_BUG_TYPES:
        return GENERIC_BUCKETS["other_bug"]
    if issue_type_name in OTHER_FEATURE_TYPES:
        return GENERIC_BUCKETS["other_feature"]
    return GENERIC_BUCKETS["other_misc"]


def _direct_dimensions(
    *,
    feature_root_id: int | None,
    feature_key: str | None,
    feature_name: str | None,
    topic_type: str | None,
    issue_type: str | None,
    families: dict[int, _FamilyInfo],
) -> tuple[str, str, str, str]:
    if feature_root_id is not None:
        return _feature_dimensions(
            feature_root_id=feature_root_id,
            feature_key=feature_key,
            feature_name=feature_name,
            families=families,
        )
    bucket_id, bucket_name = _generic_bucket(topic_type, issue_type)
    return bucket_id, bucket_name, bucket_id, bucket_name


def _person_key(account_id: str | None, display_name: str | None) -> str:
    return str(account_id or display_name or "Unknown")


def _hrworks_capacity(
    db: Session,
    *,
    account_ids: set[str],
    period_from: date,
    period_to: date,
) -> tuple[dict[tuple[str, date], float], dict[str, int]]:
    if not account_ids:
        return {}, {}
    user_rows = db.execute(
        select(JiraUser.account_id, JiraUser.id).where(JiraUser.account_id.in_(account_ids))
    ).all()
    account_to_user = {str(account_id): int(user_id) for account_id, user_id in user_rows}
    if not account_to_user:
        return {}, {}
    user_to_account = {user_id: account_id for account_id, user_id in account_to_user.items()}
    rows = db.execute(
        select(
            JiraUserMonthlyHrworksHours.jira_user_id,
            JiraUserMonthlyHrworksHours.month_start,
            JiraUserMonthlyHrworksHours.planned_working_hours,
        ).where(
            JiraUserMonthlyHrworksHours.jira_user_id.in_(set(user_to_account.keys())),
            JiraUserMonthlyHrworksHours.month_start >= period_from,
            JiraUserMonthlyHrworksHours.month_start <= period_to,
        )
    ).all()
    capacity: dict[tuple[str, date], float] = defaultdict(float)
    for jira_user_id, month_start, planned in rows:
        account_id = user_to_account.get(int(jira_user_id))
        if account_id:
            capacity[(account_id, month_start)] += float(planned or 0)
    return dict(capacity), account_to_user


def _factor_infos(
    db: Session,
    *,
    period_from: date,
    period_to: date,
) -> dict[tuple[str, date], _FactorInfo]:
    stmt = scope_monthly_topic_effort(
        select(MonthlyTopicEffortBase).where(
            MonthlyTopicEffortBase.period_month >= period_from,
            MonthlyTopicEffortBase.period_month <= period_to,
        )
    )
    rows = db.execute(stmt).scalars().all()
    account_ids = {row.user_account_id for row in rows if row.user_account_id}
    capacity, account_to_user = _hrworks_capacity(
        db,
        account_ids=set(account_ids),
        period_from=period_from,
        period_to=period_to,
    )
    totals: dict[tuple[str, date], float] = defaultdict(float)
    meta: dict[tuple[str, date], MonthlyTopicEffortBase] = {}
    for row in rows:
        bucket = _role_bucket(row.role_name)
        if not _is_dev_qa_bucket(bucket):
            continue
        key = (_person_key(row.user_account_id, row.display_name), row.period_month)
        totals[key] += float(row.direct_hours)
        meta[key] = row

    factors: dict[tuple[str, date], _FactorInfo] = {}
    for key, booked_hours in totals.items():
        person_key, period = key
        row = meta[key]
        planned = capacity.get((person_key, period), 0.0)
        target = planned * DEV_QA_CAPACITY_FACTOR
        missing = planned <= 0
        raw_factor = (target / booked_hours) if booked_hours > 0 and planned > 0 else 1.0
        factor = max(1.0, raw_factor)
        role = row.role_name or "Unknown"
        factors[key] = _FactorInfo(
            person_key=person_key,
            person=row.display_name or row.user_account_id or "Unknown",
            account_id=row.user_account_id,
            jira_user_id=account_to_user.get(person_key),
            role=role,
            role_bucket=_role_bucket(role),
            team=normalized_team_name(row.team_name),
            period=period,
            planned_hours=round(planned, 2),
            direct_booked_hours=round(booked_hours, 2),
            capacity_target_hours=round(target, 2),
            scale_factor=round(factor, 8),
            missing_hrworks=missing,
        )
    return factors


def _direct_rows(
    db: Session,
    *,
    period_from: date,
    period_to: date,
    factors: dict[tuple[str, date], _FactorInfo],
    families: dict[int, _FamilyInfo],
    feature_meta: dict[int, _FeatureMeta],
    family_meta: dict[str, _FeatureMeta],
) -> list[_AuditRow]:
    stmt = scope_monthly_topic_effort(
        select(MonthlyTopicEffortBase).where(
            MonthlyTopicEffortBase.period_month >= period_from,
            MonthlyTopicEffortBase.period_month <= period_to,
        )
    )
    out: list[_AuditRow] = []
    for row in db.execute(stmt).scalars().all():
        family_id, family_name, feature_id, feature_name = _direct_dimensions(
            feature_root_id=row.feature_root_id,
            feature_key=row.feature_key,
            feature_name=row.feature_name,
            topic_type=row.topic_type,
            issue_type=row.issue_type_name,
            families=families,
        )
        (
            family_start_date,
            family_end_date,
            family_status,
            feature_start_date,
            feature_end_date,
            feature_status,
        ) = _metadata_dimensions(
            feature_root_id=row.feature_root_id,
            family_identifier=family_id,
            feature_meta=feature_meta,
            family_meta=family_meta,
        )
        person_key = _person_key(row.user_account_id, row.display_name)
        role = row.role_name or "Unknown"
        bucket = _role_bucket(role)
        factor_info = factors.get((person_key, row.period_month))
        factor = factor_info.scale_factor if factor_info and _is_dev_qa_bucket(bucket) else 1.0
        planned = factor_info.planned_hours if factor_info and _is_dev_qa_bucket(bucket) else 0.0
        booked = float(row.direct_hours)
        out.append(
            _AuditRow(
                period=row.period_month,
                family_identifier=family_id,
                family_name=family_name,
                family_start_date=family_start_date,
                family_end_date=family_end_date,
                family_status=family_status,
                feature_identifier=feature_id,
                feature_name=feature_name,
                feature_start_date=feature_start_date,
                feature_end_date=feature_end_date,
                feature_status=feature_status,
                team=normalized_team_name(row.team_name),
                issue_identifier=row.issue_key,
                issue_name=row.summary,
                issue_type=row.issue_type_name,
                person_key=person_key,
                person=row.display_name or row.user_account_id or "Unknown",
                role=role,
                role_bucket=bucket,
                source="Jira worklog",
                booked_hours=booked,
                calculated_hours=booked * factor,
                overhead_hours=0.0,
                scale_factor=factor,
                hrworks_planned_hours=planned,
            )
        )
    return out


def _overhead_rows(
    db: Session,
    *,
    period_from: date,
    period_to: date,
    families: dict[int, _FamilyInfo],
    feature_meta: dict[int, _FeatureMeta],
    family_meta: dict[str, _FeatureMeta],
) -> list[_AuditRow]:
    stmt = scope_monthly_allocated_effort(
        select(
            MonthlyAllocatedEffort,
            JiraIssue.issue_type_name,
            JiraIssue.summary,
        )
        .outerjoin(JiraIssue, JiraIssue.id == MonthlyAllocatedEffort.issue_id)
        .where(
            MonthlyAllocatedEffort.period_month >= period_from,
            MonthlyAllocatedEffort.period_month <= period_to,
            MonthlyAllocatedEffort.allocation_kind == "indirect_allocated",
        )
    )
    out: list[_AuditRow] = []
    for row, issue_type, issue_summary in db.execute(stmt).all():
        family_id, family_name, feature_id, feature_name = _direct_dimensions(
            feature_root_id=row.feature_root_id,
            feature_key=row.feature_key,
            feature_name=row.feature_name,
            topic_type=row.topic_type,
            issue_type=issue_type,
            families=families,
        )
        (
            family_start_date,
            family_end_date,
            family_status,
            feature_start_date,
            feature_end_date,
            feature_status,
        ) = _metadata_dimensions(
            feature_root_id=row.feature_root_id,
            family_identifier=family_id,
            feature_meta=feature_meta,
            family_meta=family_meta,
        )
        role = row.source_role_name or "Unknown"
        hours = float(row.hours)
        out.append(
            _AuditRow(
                period=row.period_month,
                family_identifier=family_id,
                family_name=family_name,
                family_start_date=family_start_date,
                family_end_date=family_end_date,
                family_status=family_status,
                feature_identifier=feature_id,
                feature_name=feature_name,
                feature_start_date=feature_start_date,
                feature_end_date=feature_end_date,
                feature_status=feature_status,
                team=normalized_team_name(row.team_name),
                issue_identifier=row.issue_key or "unknown",
                issue_name=issue_summary,
                issue_type=issue_type,
                person_key=row.source_user_email,
                person=row.source_display_name,
                role=role,
                role_bucket=_role_bucket(role),
                source="Allocated overhead",
                booked_hours=0.0,
                calculated_hours=hours,
                overhead_hours=hours,
                scale_factor=1.0,
                hrworks_planned_hours=0.0,
            )
        )
    return out


def _filtered_rows(
    rows: list[_AuditRow],
    *,
    team: str | None = None,
    role: str | None = None,
    family_id: str | None = None,
    feature_key: str | None = None,
    period_month: date | None = None,
    issue_key: str | None = None,
) -> list[_AuditRow]:
    team_filter = normalized_team_name(team) if team else None
    role_filter = _role_bucket(role) if role else None
    return [
        row
        for row in rows
        if (not team_filter or row.team == team_filter)
        and (not role_filter or row.role_bucket == role_filter or row.role == role)
        and (not family_id or row.family_identifier == str(family_id))
        and (not feature_key or row.feature_identifier == feature_key)
        and (not period_month or row.period == period_month)
        and (not issue_key or row.issue_identifier == issue_key)
    ]


def _audit_rows(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
) -> tuple[list[_AuditRow], dict[tuple[str, date], _FactorInfo], list[str]]:
    period_from, period_to = _monthly_period_bounds(date_from, date_to)
    factors = _factor_infos(db, period_from=period_from, period_to=period_to)
    families = _family_map(db)
    feature_meta = _feature_meta_map(db)
    family_meta = _family_meta_map(db, feature_meta=feature_meta)
    rows = [
        *_direct_rows(
            db,
            period_from=period_from,
            period_to=period_to,
            factors=factors,
            families=families,
            feature_meta=feature_meta,
            family_meta=family_meta,
        ),
        *_overhead_rows(
            db,
            period_from=period_from,
            period_to=period_to,
            families=families,
            feature_meta=feature_meta,
            family_meta=family_meta,
        ),
    ]
    return rows, factors, [month.isoformat() for month in _month_starts(period_from, period_to)]


def _round(value: float) -> float:
    return round(value, 2)


def _leaderboard(rows: list[_AuditRow], periods: list[str]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str | None], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.family_identifier,
            row.family_name,
            row.family_start_date,
            row.family_end_date,
            row.family_status,
            row.feature_identifier,
            row.feature_name,
            row.feature_start_date,
            row.feature_end_date,
            row.feature_status,
            row.team,
        )
        entry = grouped.setdefault(
            key,
            {
                "booked_hours": 0.0,
                "calculated_hours": 0.0,
                "overhead_hours": 0.0,
                "monthly": {
                    period: {"booked": 0.0, "calculated": 0.0, "overhead": 0.0}
                    for period in periods
                },
            },
        )
        period = row.period.isoformat()
        entry["booked_hours"] += row.booked_hours
        entry["calculated_hours"] += row.calculated_hours
        entry["overhead_hours"] += row.overhead_hours
        entry["monthly"][period]["booked"] += row.booked_hours
        entry["monthly"][period]["calculated"] += row.calculated_hours
        entry["monthly"][period]["overhead"] += row.overhead_hours

    ranked = []
    for index, (key, values) in enumerate(
        sorted(grouped.items(), key=lambda item: -float(item[1]["calculated_hours"])),
        start=1,
    ):
        (
            family_id,
            family_name,
            family_start_date,
            family_end_date,
            family_status,
            feature_id,
            feature_name,
            feature_start_date,
            feature_end_date,
            feature_status,
            team,
        ) = key
        ranked.append(
            {
                "rank": index,
                "family_identifier": family_id,
                "family_name": family_name,
                "family_start_date": family_start_date,
                "family_end_date": family_end_date,
                "family_status": family_status,
                "feature_identifier": feature_id,
                "feature_name": feature_name,
                "feature_start_date": feature_start_date,
                "feature_end_date": feature_end_date,
                "feature_status": feature_status,
                "team": team,
                "booked_hours": _round(values["booked_hours"]),
                "calculated_hours": _round(values["calculated_hours"]),
                "overhead_hours": _round(values["overhead_hours"]),
                "monthly": {
                    period: {metric: _round(amount) for metric, amount in metrics.items()}
                    for period, metrics in values["monthly"].items()
                },
            }
        )
    return ranked


def _series(rows: list[_AuditRow]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str | None, str], dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    for row in rows:
        key = (
            row.family_identifier,
            row.family_name,
            row.feature_identifier,
            row.feature_name,
            row.team,
            row.period.isoformat(),
        )
        grouped[key]["booked_hours"] += row.booked_hours
        grouped[key]["calculated_hours"] += row.calculated_hours
        grouped[key]["overhead_hours"] += row.overhead_hours
        grouped[key][row.role_bucket] += row.calculated_hours
    return [
        {
            "family_identifier": family_id,
            "family_name": family_name,
            "feature_identifier": feature_id,
            "feature_name": feature_name,
            "team": team,
            "period": period,
            **{metric: _round(amount) for metric, amount in values.items()},
        }
        for (family_id, family_name, feature_id, feature_name, team, period), values in sorted(
            grouped.items(), key=lambda item: item[0]
        )
    ]


def feature_investment_audit(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
    team: str | None = None,
    role: str | None = None,
    family_id: str | None = None,
    feature_key: str | None = None,
) -> AnalyticsReportResponse:
    rows, factors, periods = _audit_rows(db, date_from=date_from, date_to=date_to)
    visible = _filtered_rows(
        rows,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
    )
    available_teams = sorted({row.team for row in rows if row.team})
    available_roles = sorted({row.role for row in rows if row.role})
    available_families = sorted(
        {row.family_identifier: row.family_name for row in rows}.items(),
        key=lambda item: item[1].lower(),
    )
    return AnalyticsReportResponse(
        filters={
            "from": periods[0] if periods else None,
            "to": _month_end(date.fromisoformat(periods[-1])).isoformat() if periods else None,
            "team": team,
            "role": role,
            "family_id": family_id,
            "feature_key": feature_key,
            "periods": periods,
            "available_teams": available_teams,
            "available_roles": available_roles,
            "available_families": [
                {"identifier": identifier, "name": name} for identifier, name in available_families
            ],
        },
        summary={
            "features": len({row.feature_identifier for row in visible}),
            "families": len({row.family_identifier for row in visible}),
            "booked_hours": _round(sum(row.booked_hours for row in visible)),
            "calculated_hours": _round(sum(row.calculated_hours for row in visible)),
            "overhead_hours": _round(sum(row.overhead_hours for row in visible)),
            "missing_hrworks_people": sorted(
                {factor.person for factor in factors.values() if factor.missing_hrworks}
            ),
        },
        table=_leaderboard(visible, periods),
        series=_series(visible),
    )


def feature_investment_audit_issues(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
    period_month: date | None = None,
    team: str | None = None,
    role: str | None = None,
    family_id: str | None = None,
    feature_key: str | None = None,
) -> AnalyticsReportResponse:
    rows, _factors, _periods = _audit_rows(db, date_from=date_from, date_to=date_to)
    visible = _filtered_rows(
        rows,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
        period_month=period_month,
    )
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in visible:
        key = (
            row.issue_identifier,
            row.family_name,
            row.feature_identifier,
            row.feature_name,
        )
        entry = grouped.setdefault(
            key,
            {
                "issue_name": row.issue_name,
                "issue_type": row.issue_type,
                "booked_hours": 0.0,
                "calculated_hours": 0.0,
                "overhead_hours": 0.0,
            },
        )
        entry["issue_name"] = entry["issue_name"] or row.issue_name
        entry["issue_type"] = entry["issue_type"] or row.issue_type
        entry["booked_hours"] += row.booked_hours
        entry["calculated_hours"] += row.calculated_hours
        entry["overhead_hours"] += row.overhead_hours
    table = [
        {
            "issue_identifier": issue_key,
            "issue_name": values["issue_name"],
            "issue_type": values["issue_type"],
            "family_name": family_name,
            "feature_identifier": feature_id,
            "feature_name": feature_name,
            "booked_hours": _round(values["booked_hours"]),
            "calculated_hours": _round(values["calculated_hours"]),
            "overhead_hours": _round(values["overhead_hours"]),
        }
        for (issue_key, family_name, feature_id, feature_name), values in sorted(
            grouped.items(), key=lambda item: item[0][0]
        )
    ]
    return AnalyticsReportResponse(filters={}, table=table)


def feature_investment_audit_worklogs(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
    issue_key: str,
    period_month: date | None = None,
) -> AnalyticsReportResponse:
    rows, factors, _periods = _audit_rows(db, date_from=date_from, date_to=date_to)
    visible = _filtered_rows(rows, issue_key=issue_key, period_month=period_month)
    direct_periods = {row.period for row in visible if row.booked_hours > 0}
    issue = db.execute(
        select(JiraIssue).where(JiraIssue.key == issue_key).limit(1)
    ).scalar_one_or_none()
    worklog_rows: list[dict[str, Any]] = []
    if issue is not None:
        wl_stmt = apply_worklog_issue_scope(
            select(JiraWorklog).where(JiraWorklog.issue_id == issue.id)
        )
        if period_month:
            wl_stmt = wl_stmt.where(JiraWorklog.started_at >= period_month).where(
                JiraWorklog.started_at < _next_month_start(period_month)
            )
        worklogs = db.execute(wl_stmt).scalars().all()
        direct_by_account_period = {
            (row.person_key, row.period): row for row in visible if row.booked_hours > 0
        }
        for wl in worklogs:
            if wl.started_at is None:
                continue
            month = date(wl.started_at.year, wl.started_at.month, 1)
            if direct_periods and month not in direct_periods:
                continue
            person_key = _person_key(wl.author_account_id, wl.author_display_name)
            factor = factors.get((person_key, month))
            matching = direct_by_account_period.get((person_key, month))
            raw_hours = float(wl.time_spent_seconds) / 3600
            scale = (
                factor.scale_factor
                if factor and matching and _is_dev_qa_bucket(matching.role_bucket)
                else 1
            )
            worklog_rows.append(
                {
                    "period": month.isoformat(),
                    "person": wl.author_display_name or "Unknown",
                    "role": matching.role if matching else None,
                    "issue_key": issue_key,
                    "worklog_id": wl.jira_worklog_id,
                    "worklog_date": wl.started_at.date().isoformat(),
                    "source": "Jira worklog",
                    "booked_hours": _round(raw_hours),
                    "calculated_hours": _round(raw_hours * scale),
                    "overhead_hours": 0.0,
                    "scale_factor": round(scale, 4),
                    "hrworks_planned_hours": factor.planned_hours if factor else 0.0,
                }
            )
    for row in visible:
        if row.overhead_hours <= 0:
            continue
        worklog_rows.append(
            {
                "period": row.period.isoformat(),
                "person": row.person,
                "role": row.role,
                "issue_key": row.issue_identifier,
                "worklog_id": None,
                "worklog_date": None,
                "source": row.source,
                "booked_hours": 0.0,
                "calculated_hours": _round(row.calculated_hours),
                "overhead_hours": _round(row.overhead_hours),
                "scale_factor": 1.0,
                "hrworks_planned_hours": 0.0,
            }
        )
    return AnalyticsReportResponse(
        filters={"issue_key": issue_key},
        table=sorted(
            worklog_rows,
            key=lambda item: (item["period"], item["person"], item["source"]),
        ),
    )


def _append_sheet(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            52,
        )


def _period_text(periods: list[str]) -> str:
    if not periods:
        return "Kein Monatszeitraum im Export enthalten."
    ordered = sorted(periods)
    return f"{_month_label(ordered[0])} bis {_month_label(ordered[-1])}"


def _filter_text(
    *,
    date_from: date | None,
    date_to: date | None,
    team: str | None,
    role: str | None,
    family_id: str | None,
    feature_key: str | None,
) -> str:
    filters = [
        f"Datum von: {date_from.isoformat() if date_from else 'nicht gesetzt'}",
        f"Datum bis: {date_to.isoformat() if date_to else 'nicht gesetzt'}",
        f"Team: {team or 'alle'}",
        f"Rolle: {role or 'alle'}",
        f"Feature Family: {family_id or 'alle'}",
        f"Feature: {feature_key or 'alle'}",
    ]
    return "; ".join(filters)


def _metric_definition_text() -> str:
    return (
        "Der Export unterscheidet bewusst zwischen gebuchten Stunden, berechneten "
        "Stunden und Overhead-Stunden.\n\n"
        "Gebuchte Stunden (Booked / Actual): Das sind die tatsaechlich in Jira "
        "erfassten Worklog-Stunden auf konkreten Issues. Sie sind die belegbare "
        "Ist-Basis des Reports und zeigen, was direkt auf Tickets gebucht wurde.\n\n"
        "Berechnete Stunden (Calculated): Diese Kennzahl korrigiert Unterbuchung "
        "bei Developer- und QA-Rollen. Fuer jede Person und jeden Monat werden "
        "alle direkt gebuchten Jira-Stunden betrachtet, also Features, Bugs, "
        "Improvements, technische Arbeit und Support. Dann wird die HRWorks-"
        "Planverfuegbarkeit herangezogen.\n\n"
        "Formel fuer Developer und QA: Calculated hours = booked hours * scale "
        "factor. Der scale factor ist max(1, HRWorks planned hours * 80% / alle "
        "direkt gebuchten Jira-Stunden der Person im Monat). Die 80% bedeuten: "
        "20% der Dev/QA-Verfuegbarkeit werden pauschal als allgemeine, nicht "
        "direkt featurebezogene Arbeit angenommen. Der Faktor wird nie kleiner "
        "als 1; gebuchte Jira-Stunden werden also nie nach unten korrigiert.\n\n"
        "Fuer UX und andere nicht skalierte direkte Rollen gilt: Calculated hours "
        "= booked hours. UX wird nicht auf HRWorks hochgerechnet, sondern bleibt "
        "bei den tatsaechlich gebuchten Jira-Stunden.\n\n"
        "Overhead-Stunden: Overhead sind Stunden von Rollen, deren Arbeit als "
        "steuernde, koordinierende, produktbezogene oder architektonische "
        "Kapazitaet verstanden wird, zum Beispiel Product Owner, Product Manager, "
        "System Architect, Head of Development oder vergleichbare konfigurierte "
        "Rollen. Welche Personen/Rollen Overhead sind und mit welchem Prozentsatz "
        "sie verteilt werden, kommt aus der Allocation-Konfiguration bzw. dem "
        "Rule Snapshot der monatlichen Allocation. Dieser Prozentsatz ist nicht "
        "hart im Export verdrahtet und kann je Rolle oder Person unterschiedlich "
        "sein.\n\n"
        "Der allocatable Anteil der HRWorks-Verfuegbarkeit einer Overhead-Person "
        "wird proportional auf die Themen verteilt, auf denen direkte Arbeit "
        "stattgefunden hat. Beispiel: 160 HRWorks-Stunden mit 50% allocatable "
        "ergeben 80 verteilbare Overhead-Stunden. Entfallen 60% der direkten "
        "Arbeit auf Feature A und 40% auf Bugs, erhaelt Feature A 48 Stunden "
        "Overhead und der Bug-Bucket 32 Stunden Overhead. Nicht allocatable "
        "Anteile bleiben shared bzw. nicht direkt zugeordnet.\n\n"
        "Wichtig: Overhead wird nicht nur auf PMGT Features verteilt. Wenn direkte "
        "Arbeit in generischen Buckets liegt, erhalten auch Other bug, Other "
        "feature und Other misc anteiligen Overhead. Dadurch zeigt der Export "
        "nicht nur Feature-Investment, sondern auch Investment in Bugs, "
        "Improvements, technische Arbeit und Support inklusive Steuerungsaufwand.\n\n"
        "Calculated Total = Calculated hours + Overhead hours. Diese Kennzahl ist "
        "die wichtigste Investment-Sicht, weil sie hochgerechnete Dev/QA-Arbeit, "
        "nicht skalierte direkte Arbeit und zugeordneten Overhead zusammenfuehrt.\n\n"
        "Actual Total = Booked hours + Overhead hours. Diese Kennzahl ist die "
        "konservativere Sicht, weil sie direkte Arbeit nicht hochrechnet.\n\n"
        "Sonderfaelle: Fehlende HRWorks-Daten fuehren zu scale factor 1. Wenn eine "
        "Person keine direkten Jira-Stunden im Monat gebucht hat, kann ihre "
        "Kapazitaet nicht proportional ueber direkte Arbeit verteilt werden. "
        "Fehlende Feature-Zuordnungen landen in generischen Buckets. Nicht "
        "eindeutige Issue-Typen werden konservativ als Other misc klassifiziert. "
        "Quellzeilen werden so summiert, wie sie in den vorbereiteten Reporting-"
        "Tabellen vorliegen; der Export fuehrt keine eigene technische "
        "Deduplizierung durch."
    )


def _methodology_rows(
    *,
    date_from: date | None,
    date_to: date | None,
    team: str | None,
    role: str | None,
    family_id: str | None,
    feature_key: str | None,
    periods: list[str],
) -> list[tuple[str, str]]:
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    effective_date_to = (
        _month_end(date.fromisoformat(sorted(periods)[-1])) if periods else date_to
    )
    filters = _filter_text(
        date_from=date_from,
        date_to=effective_date_to,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
    )
    return [
        (
            "1. Zweck des Exports",
            "Der Export beantwortet, wie viel Investment in PMGT Features, Feature Families und "
            "generische Buckets geflossen ist. Er zeigt Ist-Stunden, berechnete Stunden und "
            "zugeordneten Overhead, damit Feature- und Nicht-Feature-Arbeit gemeinsam "
            f"abgestimmt werden koennen. Gueltiger Monatszeitraum: {_period_text(periods)}. "
            f"Datenstand: Zeitpunkt der Exporterzeugung {generated_at} UTC.",
        ),
        (
            "2. Datenquellen",
            "Verwendet werden MonthlyTopicEffortBase fuer direkte Jira-Worklog-Stunden je "
            "Issue, Person, Rolle, Team, Topic und Monat; MonthlyAllocatedEffort fuer "
            "indirekt zugeordneten Overhead; JiraUserMonthlyHrworksHours fuer HRWorks "
            "Planstunden; JiraFeatureRoot, JiraFeatureFamily und JiraFeatureFamilyMember "
            "fuer Feature- und Family-Zuordnung; JiraIssue und JiraIssueDetail fuer "
            "Issue-Key, Summary, Typ, Status, Start- und Enddaten.",
        ),
        (
            "3. Filterlogik",
            "Beruecksichtigt werden Monatszeilen im ausgewaehlten Zeitraum und innerhalb "
            "des aktiven Jira-Analytics-Projektscopes. Direkte Zeilen stammen aus "
            "MonthlyTopicEffortBase; Overhead-Zeilen aus MonthlyAllocatedEffort mit "
            "allocation_kind = indirect_allocated. Angewendete Exportfilter: "
            f"{filters}. Ausgeschlossen werden Zeilen ausserhalb des Zeitraums, "
            "ausserhalb der Projekt-Scope-Regeln oder ausserhalb der gesetzten "
            "Team-, Rollen-, Family- oder Feature-Filter.",
        ),
        (
            "4. Klassifizierungslogik",
            "Wenn eine feature_root_id vorhanden ist, wird die Zeile dem PMGT Feature und "
            "der aktiven Feature Family zugeordnet. Ohne Feature wird anhand von "
            "topic_type und Jira-Issue-Typ klassifiziert: Bug/Problem zu Other bug, "
            "Feature-, Epic- oder Improvement-aehnliche Typen zu Other feature, "
            "Support, technische oder nicht eindeutig klassifizierbare Arbeit zu "
            "Other misc. Overhead-Zeilen nutzen dieselbe Klassifizierung wie direkte "
            "Zeilen, damit generische Buckets ebenfalls Overhead erhalten.",
        ),
        (
            "5. Aggregationslogik",
            "Die Summary-Sheets gruppieren nach Feature Family, Feature Name und Feature "
            "Status. Monats-Spalten stehen absteigend vom neuesten zum aeltesten Monat. "
            "Zeilen werden nach Total des neuesten Monats absteigend sortiert, danach "
            "nach den vorherigen Monaten. Detail-Sheets gruppieren bzw. sortieren nach "
            "Issue, Person, Rolle, Feature, Family, Team und Monat. Verwendet werden "
            "Summen, keine Durchschnitte, Mediane oder gewichteten Perzentile.",
        ),
        (
            "6. Berechnungslogik der Kennzahlen",
            _metric_definition_text(),
        ),
        (
            "7. Aktualitaet und Einschraenkungen",
            f"Export erzeugt am {generated_at} UTC. Die Werte haengen vom letzten Jira-, "
            "HRWorks- und Allocation-Rebuild ab. Unvollstaendig koennen fehlende "
            "HRWorks-Stunden, fehlende Rollen-/Team-Zuordnungen, fehlende Feature-"
            "Memberships oder nicht gepflegte Jira-Metadaten sein. Generische Buckets "
            "sind bewusst Sammelkategorien fuer Arbeit ohne PMGT Feature.",
        ),
    ]


def _add_methodology_sheet(
    workbook: Workbook,
    *,
    date_from: date | None,
    date_to: date | None,
    team: str | None,
    role: str | None,
    family_id: str | None,
    feature_key: str | None,
    periods: list[str],
) -> Any:
    sheet = workbook.create_sheet("Methodik & Definitionen", 1)
    sheet.sheet_properties.tabColor = "806000"
    sheet.append(["Feature Investment Audit - Methodik & Definitionen", None])
    sheet.append(["Bereich", "Beschreibung"])
    for area, description in _methodology_rows(
        date_from=date_from,
        date_to=date_to,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
        periods=periods,
    ):
        sheet.append([area, description])

    sheet.merge_cells("A1:B1")
    sheet.freeze_panes = "A3"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 110
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(vertical="center")
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=3, max_col=2):
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return sheet


def _role_rollup(rows: list[_AuditRow], *, family_only: bool) -> list[list[Any]]:
    grouped: dict[tuple[Any, ...], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        if family_only:
            key = (
                row.family_identifier,
                row.family_name,
                row.family_start_date,
                row.family_end_date,
                row.family_status,
                row.period.isoformat(),
                row.team,
                row.role_bucket,
            )
        else:
            key = (
                row.family_identifier,
                row.family_name,
                row.family_start_date,
                row.family_end_date,
                row.family_status,
                row.feature_identifier,
                row.feature_name,
                row.feature_start_date,
                row.feature_end_date,
                row.feature_status,
                row.period.isoformat(),
                row.team,
                row.role_bucket,
            )
        grouped[key]["booked"] += row.booked_hours
        grouped[key]["calculated"] += row.calculated_hours
        grouped[key]["overhead"] += row.overhead_hours
    return [
        [*key, _round(values["booked"]), _round(values["calculated"]), _round(values["overhead"])]
        for key, values in sorted(grouped.items(), key=lambda item: item[0])
    ]


def _month_label(period: str) -> str:
    return date.fromisoformat(period).strftime("%Y-%m")


def _summary_headers(periods: list[str], *, actual: bool) -> list[str]:
    headers = ["Feature Family", "Feature Name", "Feature Status"]
    basis = "Actual" if actual else "Calculated"
    total = "Actual Total" if actual else "Calc Total"
    for period in sorted(periods, reverse=True):
        label = _month_label(period)
        headers.extend(
            [
                f"{label} {basis}",
                f"{label} Overhead",
                f"{label} {total}",
            ]
        )
    return headers


def _summary_rollup(rows: list[_AuditRow], periods: list[str], *, actual: bool) -> list[list[Any]]:
    ordered_periods = sorted(periods, reverse=True)
    grouped: dict[tuple[str, str, str], dict[str, dict[str, float]]] = defaultdict(
        lambda: {
            period: {"direct": 0.0, "overhead": 0.0}
            for period in periods
        }
    )
    for row in rows:
        period = row.period.isoformat()
        if period not in periods:
            continue
        key = (
            row.family_name,
            row.feature_name,
            row.feature_status or "",
        )
        grouped[key][period]["direct"] += row.booked_hours if actual else (
            row.calculated_hours - row.overhead_hours
        )
        grouped[key][period]["overhead"] += row.overhead_hours

    out: list[list[Any]] = []
    for key, values in sorted(
        grouped.items(),
        key=lambda item: (
            *[
                -(item[1][period]["direct"] + item[1][period]["overhead"])
                for period in ordered_periods
            ],
            item[0],
        ),
    ):
        row: list[Any] = [*key]
        for period in ordered_periods:
            direct = values[period]["direct"]
            overhead = values[period]["overhead"]
            row.extend([_round(direct), _round(overhead), _round(direct + overhead)])
        out.append(row)
    return out


def _hrworks_planned_lookup(
    db: Session,
    *,
    person_keys: set[str],
    periods: set[date],
) -> dict[tuple[str, date], float]:
    if not person_keys or not periods:
        return {}
    users = db.execute(
        select(JiraUser).where(
            or_(
                JiraUser.account_id.in_(person_keys),
                JiraUser.email_address.in_(person_keys),
            )
        )
    ).scalars().all()
    key_by_user_id: dict[int, set[str]] = defaultdict(set)
    for user in users:
        if user.account_id in person_keys:
            key_by_user_id[user.id].add(user.account_id)
        if user.email_address in person_keys:
            key_by_user_id[user.id].add(user.email_address)
    if not key_by_user_id:
        return {}

    rows = db.execute(
        select(JiraUserMonthlyHrworksHours).where(
            JiraUserMonthlyHrworksHours.jira_user_id.in_(key_by_user_id),
            JiraUserMonthlyHrworksHours.month_start.in_(periods),
        )
    ).scalars().all()
    planned: dict[tuple[str, date], float] = {}
    for row in rows:
        for person_key in key_by_user_id.get(row.jira_user_id, set()):
            planned[(person_key, row.month_start)] = float(row.planned_working_hours)
    return planned


def _hrworks_audit_rollup(
    db: Session,
    *,
    rows: list[_AuditRow],
    factors: dict[tuple[str, date], _FactorInfo],
) -> list[list[Any]]:
    grouped: dict[tuple[str, date, str, str | None], dict[str, Any]] = defaultdict(
        lambda: {
            "person": "Unknown",
            "direct_booked_hours": 0.0,
            "overhead_hours": 0.0,
            "sources": set(),
        }
    )
    for row in rows:
        key = (row.person_key, row.period, row.role, row.team)
        entry = grouped[key]
        entry["person"] = row.person or entry["person"]
        entry["direct_booked_hours"] += row.booked_hours
        entry["overhead_hours"] += row.overhead_hours
        entry["sources"].add(row.source)

    planned = _hrworks_planned_lookup(
        db,
        person_keys={key[0] for key in grouped},
        periods={key[1] for key in grouped},
    )

    out: list[list[Any]] = []
    for (person_key, period, role, team), values in grouped.items():
        factor = factors.get((person_key, period))
        role_bucket = _role_bucket(role)
        if factor and _is_dev_qa_bucket(role_bucket):
            planned_hours = factor.planned_hours
            direct_booked_hours = factor.direct_booked_hours
            capacity_target_hours = factor.capacity_target_hours
            scale_factor = factor.scale_factor
            missing_hrworks = factor.missing_hrworks
        else:
            planned_hours = round(planned.get((person_key, period), 0.0), 2)
            direct_booked_hours = round(float(values["direct_booked_hours"]), 2)
            capacity_target_hours = 0.0
            scale_factor = 1.0
            missing_hrworks = planned_hours <= 0
        out.append(
            [
                values["person"],
                role,
                team,
                period.isoformat(),
                planned_hours,
                direct_booked_hours,
                capacity_target_hours,
                round(scale_factor, 4),
                missing_hrworks,
                _round(float(values["overhead_hours"])),
                ", ".join(sorted(values["sources"])),
            ]
        )
    return sorted(out, key=lambda item: (item[3], str(item[0]).lower(), str(item[1]).lower()))


def feature_investment_audit_xlsx(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
    team: str | None = None,
    role: str | None = None,
    family_id: str | None = None,
    feature_key: str | None = None,
) -> bytes:
    rows, factors, periods = _audit_rows(db, date_from=date_from, date_to=date_to)
    visible = _filtered_rows(
        rows,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
    )
    workbook = Workbook()
    calculated_summary_sheet = workbook.active
    calculated_summary_sheet.title = "Calculated summary"
    _append_sheet(
        calculated_summary_sheet,
        _summary_headers(periods, actual=False),
        _summary_rollup(visible, periods, actual=False),
    )
    actual_summary_sheet = workbook.create_sheet("Actual summary")
    _append_sheet(
        actual_summary_sheet,
        _summary_headers(periods, actual=True),
        _summary_rollup(visible, periods, actual=True),
    )
    issue_sheet = workbook.create_sheet("Issue detail")
    issue_sheet.title = "Issue detail"
    detail_rows = sorted(
        visible,
        key=lambda row: (row.issue_identifier, row.period, row.person.lower(), row.source),
    )
    _append_sheet(
        issue_sheet,
        [
            "Feature family identifier",
            "Feature family name",
            "Feature family start date",
            "Feature family end date",
            "Feature family status",
            "Feature identifier",
            "Feature name",
            "Feature start date",
            "Feature end date",
            "Feature status",
            "Month",
            "Team",
            "Issue identifier",
            "Issue name",
            "Issue type",
            "Person",
            "Role",
            "Source",
            "Booked hours",
            "Calculated hours",
            "Overhead hours",
            "Scale factor",
        ],
        [
            [
                row.family_identifier,
                row.family_name,
                row.family_start_date,
                row.family_end_date,
                row.family_status,
                row.feature_identifier,
                row.feature_name,
                row.feature_start_date,
                row.feature_end_date,
                row.feature_status,
                row.period.isoformat(),
                row.team,
                row.issue_identifier,
                row.issue_name,
                row.issue_type,
                row.person,
                row.role,
                row.source,
                _round(row.booked_hours),
                _round(row.calculated_hours),
                _round(row.overhead_hours),
                round(row.scale_factor, 4),
            ]
            for row in detail_rows
        ],
    )
    feature_sheet = workbook.create_sheet("Feature by month and role")
    _append_sheet(
        feature_sheet,
        [
            "Feature family identifier",
            "Feature family name",
            "Feature family start date",
            "Feature family end date",
            "Feature family status",
            "Feature identifier",
            "Feature name",
            "Feature start date",
            "Feature end date",
            "Feature status",
            "Month",
            "Team",
            "Role bucket",
            "Booked hours",
            "Calculated hours",
            "Overhead hours",
        ],
        _role_rollup(visible, family_only=False),
    )
    family_sheet = workbook.create_sheet("Family by month and role")
    _append_sheet(
        family_sheet,
        [
            "Feature family identifier",
            "Feature family name",
            "Feature family start date",
            "Feature family end date",
            "Feature family status",
            "Month",
            "Team",
            "Role bucket",
            "Booked hours",
            "Calculated hours",
            "Overhead hours",
        ],
        _role_rollup(visible, family_only=True),
    )
    audit_sheet = workbook.create_sheet("HRWorks audit")
    _append_sheet(
        audit_sheet,
        [
            "Person",
            "Role",
            "Team",
            "Month",
            "HRWorks planned hours",
            "Direct booked hours denominator",
            "Dev/QA capacity target",
            "Scale factor",
            "Missing HRWorks",
            "Overhead hours",
            "Source types",
        ],
        _hrworks_audit_rollup(db, rows=visible, factors=factors),
    )
    _add_methodology_sheet(
        workbook,
        date_from=date_from,
        date_to=date_to,
        team=team,
        role=role,
        family_id=family_id,
        feature_key=feature_key,
        periods=periods,
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
